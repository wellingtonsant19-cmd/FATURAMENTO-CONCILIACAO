import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── DE-PARA: descrição Matera → código GIS ───────────────────────────────────
DEPARA = {
    'MAXIFROTA ABASTECIMENTO': 'MX',
    'CARTAO COMBUSTIVEL': 'VCE',
    'MAXIFROTA MANUTENCAO': 'GM',
    'MAXIFROTA PRIVATE': 'MXP',
    'VEIC ELO': 'VEIC',
    'VEIC PRE PAGO': 'VPP',
    'CARTAO ALIMENTACAO': 'VAE',
    'CARTAO REFEICAO': 'VRE',
    'CARTAO YUO': 'YUO',
    'GESTAO DE COMPRAS': 'GC',
    'MULTI BENEFICIO': 'MB',
    'NUTRICASH BENEFICIO SOCIAL': 'SOC',
    'NUTRICASH FLEX': 'FLX',
    'NUTRICASH PREMIUM': 'NP',
    'VALE COMBUSTIVEL': 'VC',
    'VALE REFEICAO': 'VR',
    'NUTRICASH CORPORATE': 'COR',
    'VALE ALIMENTACAO': 'VA',
    'PEDIDO': 'PED',
    'MANUTENCAO': 'GM',
}

# ─── CORREÇÃO: VEIC em vez de VEI ─────────────────────────────────────────────
PRODUTOS_MAXIFROTA = ['GM', 'MPP', 'MX', 'MXP', 'PED', 'VCE', 'VEIC', 'VPP']
PRODUTOS_NUTRICASH = ['COR', 'FLX', 'GC', 'MB', 'NP', 'SOC', 'VA', 'VAE', 'VC', 'VCE', 'VR', 'VRE', 'YUO']

# Mapeamento GIS "VEI" → nosso código "VEIC" (o GIS grava VEI, mas o correto é VEIC)
GIS_PRODUTO_REMAP = {'VEI': 'VEIC'}


def parse_br_float(s):
    if pd.isna(s):
        return 0.0
    s = str(s).strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0


def load_gis(uploaded_files):
    frames = []
    for f in uploaded_files:
        df = pd.read_csv(f, sep=';', encoding='utf-8')
        df.columns = [c.strip() for c in df.columns]
        df = df.rename(columns={
            'Empresa': 'empresa',
            'Dt Emissao': 'data',
            'Produto': 'produto',
            'Valor Bruto': 'vbr',
            'Nao Integrado': 'nao_int',
            'Integrado': 'integrado',
        })
        df = df.dropna(subset=['empresa'])
        df['integrado'] = df['integrado'].apply(parse_br_float)
        df['nao_int'] = df['nao_int'].apply(parse_br_float)
        df['vbr'] = df['vbr'].apply(parse_br_float)
        df['data'] = pd.to_datetime(df['data'].str.strip(), dayfirst=True, errors='coerce')
        df['produto'] = df['produto'].str.strip()
        # Remapear VEI → VEIC
        df['produto'] = df['produto'].replace(GIS_PRODUTO_REMAP)
        frames.append(df)
    return pd.concat(frames, ignore_index=True)


def load_matera(uploaded_file):
    df = pd.read_csv(uploaded_file, sep=';', encoding='latin1')
    df.columns = [c.strip() for c in df.columns]
    df['nVlr_tit'] = df['nVlr_tit'].apply(parse_br_float)
    df['dDt_emissao'] = pd.to_datetime(df['dDt_emissao'].str.strip(), dayfirst=True, errors='coerce')
    df['produto_gis'] = df['sDescricao_tipo_produto_servico'].str.strip().str.upper().map(DEPARA)
    return df


def calcular_matera_pivot(matera_df):
    m = matera_df.dropna(subset=['produto_gis', 'dDt_emissao'])
    return m.groupby(['dDt_emissao', 'produto_gis'])['nVlr_tit'].sum().reset_index()


def montar_dados(gis_df, matera_pivot, produtos, empresa_filtro):
    if empresa_filtro == 'MAXIFROTA':
        gis = gis_df[gis_df['empresa'].str.upper().str.contains('MAXIFROTA')]
    else:
        gis = gis_df[~gis_df['empresa'].str.upper().str.contains('MAXIFROTA')]

    gis_piv = gis.groupby(['data', 'produto']).agg(
        integrado=('integrado', 'sum'),
        nao_int=('nao_int', 'sum')
    ).reset_index()

    datas = pd.date_range('2026-03-01', '2026-03-31', freq='D')
    rows_gis, rows_mat, rows_conc, rows_nint = {}, {}, {}, {}

    for dt in datas:
        gis_row, mat_row, conc_row, nint_row = {}, {}, {}, {}
        for prod in produtos:
            val_gis = gis_piv[(gis_piv['data'] == dt) & (gis_piv['produto'] == prod)]['integrado'].sum()
            val_nint = gis_piv[(gis_piv['data'] == dt) & (gis_piv['produto'] == prod)]['nao_int'].sum()
            val_mat = matera_pivot[(matera_pivot['dDt_emissao'] == dt) & (matera_pivot['produto_gis'] == prod)]['nVlr_tit'].sum()
            gis_row[prod] = val_gis if val_gis != 0 else None
            mat_row[prod] = val_mat if val_mat != 0 else None
            nint_row[prod] = val_nint if val_nint != 0 else None
            conc_row[prod] = round(val_gis - val_mat, 2)
        rows_gis[dt] = gis_row
        rows_mat[dt] = mat_row
        rows_conc[dt] = conc_row
        rows_nint[dt] = nint_row

    return datas, rows_gis, rows_mat, rows_conc, rows_nint


def escrever_aba(ws, produtos, datas, rows_gis, rows_mat, rows_conc, rows_nint):
    n = len(produtos)

    # ─── Estilos idênticos ao modelo Reembolso.xlsx ─────────────────────────────
    hdr1_font = Font(name='Calibri', bold=True, size=18)
    hdr2_font = Font(name='Calibri', bold=True, size=12, color='000000')
    hdr2_font_blue = Font(name='Calibri', bold=True, size=12, color='000000')
    data_font = Font(name='Calibri', bold=True, size=11)
    val_font = Font(name='Calibri', size=11, color='000000')
    tot_font = Font(name='Calibri', bold=True, size=12)

    yellow_fill = PatternFill('solid', fgColor='FFFF00')
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right')
    center_h = Alignment(horizontal='center')

    thin_side = Side(style='thin')
    medium_side = Side(style='medium')
    border_medium = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)
    border_hdr_bottom = Border(top=medium_side, bottom=medium_side)
    border_data_lr = Border(left=thin_side, right=thin_side)
    border_right_thin = Border(right=thin_side)
    border_total = Border(left=thin_side, top=thin_side, bottom=thin_side)

    NUM_FMT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* \\-??_);_(@_)'
    TOT_FMT = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* \\-??_-;_-@_-'

    # ─── Layout de colunas ──────────────────────────────────────────────────────
    col_gis_start = 2
    col_mat_start = col_gis_start + n
    col_conc_start = col_mat_start + n + 1  # +1 para col DATA separadora
    col_data2 = col_mat_start + n  # col DATA entre Matera e Conciliação
    col_nint_data = col_conc_start + n  # col DATA antes de Não Integrado
    col_nint_start = col_nint_data + 1

    # ─── Row 1: seção headers com merge ─────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16.5

    def sec_header(col_s, col_e, titulo):
        ws.merge_cells(start_row=1, start_column=col_s, end_row=1, end_column=col_e)
        c = ws.cell(1, col_s, titulo)
        c.font = hdr1_font
        c.alignment = center
        c.border = border_medium

    sec_header(col_gis_start, col_gis_start + n - 1, 'GIS')
    sec_header(col_mat_start, col_mat_start + n - 1, 'CONTAS A PAGAR - MATERA')
    sec_header(col_conc_start, col_conc_start + n - 1, 'CONCILIAÇÃO')
    sec_header(col_nint_start, col_nint_start + n - 1, 'NÃO INTEGRADO NO GIS')

    # ─── Row 2: cabeçalhos de produto ───────────────────────────────────────────
    c_data = ws.cell(2, 1, 'DATA')
    c_data.font = hdr2_font
    c_data.alignment = center_h
    c_data.border = border_hdr_bottom

    for i, prod in enumerate(produtos):
        # GIS
        c = ws.cell(2, col_gis_start + i, prod)
        c.font = hdr2_font
        c.alignment = center_h
        c.border = border_hdr_bottom
        # Matera
        c = ws.cell(2, col_mat_start + i, prod)
        c.font = hdr2_font
        c.alignment = center_h
        c.border = border_hdr_bottom
        # Conciliação
        c = ws.cell(2, col_conc_start + i, prod)
        c.font = Font(name='Calibri', bold=True, size=12)
        c.alignment = center_h
        c.border = Border(left=thin_side, bottom=medium_side)
        c.number_format = NUM_FMT
        # Não Integrado (fundo amarelo)
        c_ni = ws.cell(2, col_nint_start + i, prod)
        c_ni.font = Font(name='Calibri', bold=True, size=12)
        c_ni.alignment = center_h
        c_ni.fill = yellow_fill
        c_ni.border = border_hdr_bottom

    # DATA cols separadoras
    c_d2 = ws.cell(2, col_data2)
    c_d2.font = hdr2_font
    c_d2.alignment = center_h
    c_d2.border = border_hdr_bottom

    c_nd = ws.cell(2, col_nint_data, 'DATA')
    c_nd.font = hdr2_font
    c_nd.alignment = center_h
    c_nd.border = border_hdr_bottom

    # ─── Linhas de dados (rows 3-33) ────────────────────────────────────────────
    for idx, dt in enumerate(datas):
        row = 3 + idx
        ws.row_dimensions[row].height = 15

        c = ws.cell(row, 1, dt)
        c.number_format = 'mm-dd-yy'
        c.font = data_font
        c.alignment = center_h
        c.border = border_data_lr

        for i, prod in enumerate(produtos):
            val_g = rows_gis[dt].get(prod)
            val_m = rows_mat[dt].get(prod)
            val_n = rows_nint[dt].get(prod)

            def _write(col, val):
                cell = ws.cell(row, col)
                if val is not None and val != 0:
                    cell.value = round(val, 2)
                cell.number_format = NUM_FMT
                cell.font = val_font

            _write(col_gis_start + i, val_g)
            _write(col_mat_start + i, val_m)
            _write(col_nint_start + i, val_n)

            # Conciliação = fórmula GIS - Matera
            gc = get_column_letter(col_gis_start + i)
            mc = get_column_letter(col_mat_start + i)
            cc = ws.cell(row, col_conc_start + i)
            cc.value = f'={gc}{row}-{mc}{row}'
            cc.number_format = NUM_FMT
            cc.font = val_font

        # Última col de cada seção com borda right thin
        ws.cell(row, col_gis_start + n - 1).border = border_right_thin
        ws.cell(row, col_mat_start + n - 1).border = border_right_thin
        ws.cell(row, col_conc_start + n - 1).border = border_right_thin

        # DATA separadora (entre Matera e Conciliação)
        c2 = ws.cell(row, col_data2, f'={get_column_letter(1)}{row}')
        c2.number_format = 'mm-dd-yy'
        c2.font = val_font

        # DATA separadora (entre Conciliação e Não Integrado)
        c3 = ws.cell(row, col_nint_data, f'={get_column_letter(1)}{row}')
        c3.number_format = 'mm-dd-yy'
        c3.font = data_font
        c3.alignment = center_h
        c3.border = border_data_lr

    # ─── Linha TOTAL (row 34) ───────────────────────────────────────────────────
    tot_row = 34
    ws.row_dimensions[tot_row].height = 15.75
    c_tot = ws.cell(tot_row, 1, 'TOTAL')
    c_tot.font = tot_font
    c_tot.alignment = center_h
    c_tot.border = border_total
    c_tot.number_format = TOT_FMT

    for i in range(n):
        for base in [col_gis_start, col_mat_start, col_nint_start]:
            col_l = get_column_letter(base + i)
            c = ws.cell(tot_row, base + i, f'=SUM({col_l}3:{col_l}33)')
            c.number_format = TOT_FMT
            c.font = tot_font
            c.alignment = center_h
            c.border = border_total
        col_l = get_column_letter(col_conc_start + i)
        c = ws.cell(tot_row, col_conc_start + i, f'=SUM({col_l}3:{col_l}33)')
        c.number_format = TOT_FMT
        c.font = tot_font
        c.alignment = center_h
        c.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # ─── Larguras de coluna (idênticas ao modelo) ──────────────────────────────
    ws.column_dimensions['A'].width = 17.3
    for base in [col_gis_start, col_mat_start]:
        for i in range(n):
            cl = get_column_letter(base + i)
            ws.column_dimensions[cl].width = 20 if i == 0 else 13
    for i in range(n):
        cl = get_column_letter(col_conc_start + i)
        ws.column_dimensions[cl].width = 21 if i == 0 else 13
    ws.column_dimensions[get_column_letter(col_data2)].width = 15
    ws.column_dimensions[get_column_letter(col_nint_data)].width = 13
    for i in range(n):
        cl = get_column_letter(col_nint_start + i)
        ws.column_dimensions[cl].width = 23 if i == 0 else 13


def gerar_excel(gis_df, nc_piv, mx_piv):
    wb = Workbook()
    wb.remove(wb.active)

    ws_mx = wb.create_sheet('MAXIFROTA')
    datas, rg, rm, rc, rn = montar_dados(gis_df, mx_piv, PRODUTOS_MAXIFROTA, 'MAXIFROTA')
    escrever_aba(ws_mx, PRODUTOS_MAXIFROTA, datas, rg, rm, rc, rn)

    ws_nc = wb.create_sheet('NUTRICASH')
    datas, rg, rm, rc, rn = montar_dados(gis_df, nc_piv, PRODUTOS_NUTRICASH, 'NUTRICASH')
    escrever_aba(ws_nc, PRODUTOS_NUTRICASH, datas, rg, rm, rc, rn)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_preview_df(datas, rows_gis, rows_mat, rows_conc, rows_nint, produtos):
    """Monta um DataFrame para visualização em tela."""
    records = []
    for dt in datas:
        row = {'DATA': dt.strftime('%d/%m/%Y')}
        for prod in produtos:
            row[f'GIS_{prod}'] = rows_gis[dt].get(prod) or 0
            row[f'MAT_{prod}'] = rows_mat[dt].get(prod) or 0
            row[f'CONC_{prod}'] = rows_conc[dt].get(prod, 0)
            row[f'NINT_{prod}'] = rows_nint[dt].get(prod) or 0
        records.append(row)
    # Linha TOTAL
    total = {'DATA': 'TOTAL'}
    for prod in produtos:
        for pfx in ['GIS', 'MAT', 'CONC', 'NINT']:
            total[f'{pfx}_{prod}'] = sum(r[f'{pfx}_{prod}'] for r in records)
    records.append(total)
    return pd.DataFrame(records)


# ─── INTERFACE STREAMLIT ───────────────────────────────────────────────────────
st.set_page_config(page_title='Conciliação GIS x Matera', layout='wide')
st.title('📊 Conciliação GIS x Matera')
st.markdown('Faça upload dos arquivos abaixo e clique em **Gerar Reembolso**.')

with st.sidebar:
    st.header('📂 Upload dos arquivos')
    gis_files = st.file_uploader(
        'Arquivos GIS (Fechamento Credenciado) — pode selecionar vários',
        type='csv', accept_multiple_files=True, key='gis'
    )
    nc_file = st.file_uploader('Matera NC (RLTITGER...NC.csv)', type='csv', key='nc')
    mx_file = st.file_uploader('Matera MX (RLTITGER...MX.csv)', type='csv', key='mx')

if gis_files and nc_file and mx_file:
    if st.button('🚀 Gerar Reembolso', type='primary'):
        with st.spinner('Processando...'):
            try:
                gis_df = load_gis(gis_files)
                nc_df = load_matera(nc_file)
                mx_df = load_matera(mx_file)
                nc_piv = calcular_matera_pivot(nc_df)
                mx_piv = calcular_matera_pivot(mx_df)
                buf = gerar_excel(gis_df, nc_piv, mx_piv)

                st.success('✅ Planilha gerada com sucesso!')

                # ─── VISUALIZAÇÃO EM TELA ──────────────────────────────────────
                tab_mx, tab_nc, tab_div = st.tabs(['📋 MAXIFROTA', '📋 NUTRICASH', '⚠️ Divergências'])

                # --- Aba MAXIFROTA ---
                with tab_mx:
                    datas_mx, rg_mx, rm_mx, rc_mx, rn_mx = montar_dados(gis_df, mx_piv, PRODUTOS_MAXIFROTA, 'MAXIFROTA')
                    df_mx = build_preview_df(datas_mx, rg_mx, rm_mx, rc_mx, rn_mx, PRODUTOS_MAXIFROTA)

                    st.subheader('GIS')
                    cols_gis = ['DATA'] + [f'GIS_{p}' for p in PRODUTOS_MAXIFROTA]
                    st.dataframe(df_mx[cols_gis].rename(columns={f'GIS_{p}': p for p in PRODUTOS_MAXIFROTA}), use_container_width=True, hide_index=True)

                    st.subheader('Contas a Pagar - Matera')
                    cols_mat = ['DATA'] + [f'MAT_{p}' for p in PRODUTOS_MAXIFROTA]
                    st.dataframe(df_mx[cols_mat].rename(columns={f'MAT_{p}': p for p in PRODUTOS_MAXIFROTA}), use_container_width=True, hide_index=True)

                    st.subheader('Conciliação (GIS - Matera)')
                    cols_conc = ['DATA'] + [f'CONC_{p}' for p in PRODUTOS_MAXIFROTA]
                    st.dataframe(df_mx[cols_conc].rename(columns={f'CONC_{p}': p for p in PRODUTOS_MAXIFROTA}), use_container_width=True, hide_index=True)

                    st.subheader('Não Integrado no GIS')
                    cols_nint = ['DATA'] + [f'NINT_{p}' for p in PRODUTOS_MAXIFROTA]
                    st.dataframe(df_mx[cols_nint].rename(columns={f'NINT_{p}': p for p in PRODUTOS_MAXIFROTA}), use_container_width=True, hide_index=True)

                # --- Aba NUTRICASH ---
                with tab_nc:
                    datas_nc, rg_nc, rm_nc, rc_nc, rn_nc = montar_dados(gis_df, nc_piv, PRODUTOS_NUTRICASH, 'NUTRICASH')
                    df_nc = build_preview_df(datas_nc, rg_nc, rm_nc, rc_nc, rn_nc, PRODUTOS_NUTRICASH)

                    st.subheader('GIS')
                    cols_gis = ['DATA'] + [f'GIS_{p}' for p in PRODUTOS_NUTRICASH]
                    st.dataframe(df_nc[cols_gis].rename(columns={f'GIS_{p}': p for p in PRODUTOS_NUTRICASH}), use_container_width=True, hide_index=True)

                    st.subheader('Contas a Pagar - Matera')
                    cols_mat = ['DATA'] + [f'MAT_{p}' for p in PRODUTOS_NUTRICASH]
                    st.dataframe(df_nc[cols_mat].rename(columns={f'MAT_{p}': p for p in PRODUTOS_NUTRICASH}), use_container_width=True, hide_index=True)

                    st.subheader('Conciliação (GIS - Matera)')
                    cols_conc = ['DATA'] + [f'CONC_{p}' for p in PRODUTOS_NUTRICASH]
                    st.dataframe(df_nc[cols_conc].rename(columns={f'CONC_{p}': p for p in PRODUTOS_NUTRICASH}), use_container_width=True, hide_index=True)

                    st.subheader('Não Integrado no GIS')
                    cols_nint = ['DATA'] + [f'NINT_{p}' for p in PRODUTOS_NUTRICASH]
                    st.dataframe(df_nc[cols_nint].rename(columns={f'NINT_{p}': p for p in PRODUTOS_NUTRICASH}), use_container_width=True, hide_index=True)

                # --- Aba Divergências ---
                with tab_div:
                    st.subheader('⚠️ Divergências encontradas (Conciliação ≠ 0)')
                    divergencias = []
                    for aba, produtos, piv, emp in [
                        ('MAXIFROTA', PRODUTOS_MAXIFROTA, mx_piv, 'MAXIFROTA'),
                        ('NUTRICASH', PRODUTOS_NUTRICASH, nc_piv, 'NUTRICASH'),
                    ]:
                        _, rg, rm, rc, _ = montar_dados(gis_df, piv, produtos, emp)
                        for dt, prods in rc.items():
                            for prod, diff in prods.items():
                                if diff != 0:
                                    divergencias.append({
                                        'Aba': aba,
                                        'Data': dt.strftime('%d/%m/%Y'),
                                        'Produto': prod,
                                        'GIS': rg[dt].get(prod) or 0,
                                        'Matera': rm[dt].get(prod) or 0,
                                        'Diferença': diff,
                                    })

                    if divergencias:
                        df_div = pd.DataFrame(divergencias)
                        st.dataframe(df_div, use_container_width=True, hide_index=True)
                    else:
                        st.info('Nenhuma divergência encontrada. Tudo conciliado!')

                # ─── BOTÃO DOWNLOAD ────────────────────────────────────────────
                st.divider()
                st.download_button(
                    label='⬇️ Baixar Reembolso.xlsx',
                    data=buf,
                    file_name='Reembolso.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    type='primary',
                )

            except Exception as e:
                st.error(f'Erro ao processar: {e}')
                import traceback
                st.code(traceback.format_exc())
else:
    st.info('👈 Faça upload de todos os arquivos na barra lateral para começar.')
